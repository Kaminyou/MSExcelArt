import argparse
import cv2
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from src.utils import rgb2hex

if __name__ == "__main__":

    parser = argparse.ArgumentParser()

    parser.add_argument('--image', type=str, help="path to the input image")
    parser.add_argument('--reduce', type=float, default=1, help="to reduce size =>(h/reduce, w/reduce)")
    parser.add_argument('--sheet', type=str, default="art", help="sheet name")
    parser.add_argument('--output', type=str, default="output.xlsx", help="output path with file name")
    
    args = parser.parse_args()

    img = cv2.imread(args.image)
    h, w = img.shape[:2]
    print(f"=========== Input image size = ({h}, {w}) ===========")
    h, w = int(h / args.reduce), int(w / args.reduce)
    print(f"=========== Output size = ({h}, {w}) ===========")
    img = cv2.resize(img, (h, h), interpolation=cv2.INTER_AREA)
    img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

    wb = Workbook() 
    ws = wb.active
    ws.title = args.sheet

    for i, row in enumerate(ws[f"{get_column_letter(1)}1:{get_column_letter(w)}{str(h)}"]):
        for j, cell in enumerate(row):
            print(f"[{i:5d},{j:5d}] Processing ...", end="\r")
            r, g, b = img[i, j]
            colorhex = rgb2hex(r,g,b)
            fill_gen = PatternFill(fill_type='solid', start_color=colorhex, end_color=colorhex)
            cell.fill = fill_gen

    for i in range(1, (w+1)):
        ws.column_dimensions[get_column_letter(i)].width = 1
        
    for i in range(1, (h+1)):
        ws.row_dimensions[i].height = 5

    wb.save(args.output)
    wb.close()
    print("=========== Finish ===========")